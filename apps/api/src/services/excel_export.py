"""Genera un archivo Excel (.xlsx) del sugerido filtrado, con las columnas elegidas.

Usa openpyxl. Aplica formato chileno: CLP sin decimales con miles en punto.
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import Sugerido

# Etiquetas legibles para las cabeceras del Excel.
LABELS: dict[str, str] = {
    "producto": "Producto",
    "descripcion": "Descripcion",
    "clasificacion_abc": "ABC",
    "nombre_sucursal": "Sucursal",
    "sucursal_id": "ID Sucursal",
    "proveedor": "Proveedor",
    "filtro1_final": "Marca",
    "tipo_origen": "Tipo Origen",
    "unidad_medida": "Unidad",
    "lead_time_dias": "Lead Time (dias)",
    "lt_efectivo": "LT Efectivo",
    "lt_origen": "Origen LT",
    "abastece_cd": "Abastece CD",
    "prioridad_cd": "Prioridad CD",
    "demanda_mensual": "Demanda Mensual",
    "demanda_diaria": "Demanda Diaria",
    "desv_std_mensual": "Desv Std Mensual",
    "stock_seguridad": "Stock Seguridad",
    "punto_de_pedido": "Punto de Pedido",
    "costo_unitario": "Costo Unitario",
    "pedir": "Pedir",
    "stock_activo_suc": "Stock Activo",
    "stock_en_transito_suc": "Stock en Transito",
    "stock_en_cd": "Stock en CD",
    "sugerido_traslado": "Sugerido Traslado",
    "sugerido_compra_neto": "Sugerido Compra Neto",
    "total_sugerido_suc": "Total Sugerido",
    "total_valor_sugerido_clp": "Valor Total CLP",
}

# Columnas por defecto si el cliente no especifica.
DEFAULT_COLUMNS = [
    "producto", "descripcion", "clasificacion_abc", "nombre_sucursal",
    "proveedor", "total_sugerido_suc", "total_valor_sugerido_clp",
]

CLP_COLUMNS = {"total_valor_sugerido_clp", "costo_unitario"}
HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def generar_excel(rows: list[Sugerido], columnas: list[str] | None) -> bytes:
    cols = [c for c in (columnas or []) if c in LABELS] or DEFAULT_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = "Sugerido"

    # Cabecera.
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=LABELS.get(col, col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos.
    for i, row in enumerate(rows, start=2):
        for j, col in enumerate(cols, start=1):
            value = getattr(row, col, None)
            cell = ws.cell(row=i, column=j, value=value)
            if col in CLP_COLUMNS and isinstance(value, (int, float)):
                cell.number_format = '"$"#,##0'
            elif isinstance(value, float):
                cell.number_format = "#,##0.00"

    # Ancho de columnas aproximado.
    for j, col in enumerate(cols, start=1):
        width = max(12, min(40, len(LABELS.get(col, col)) + 4))
        ws.column_dimensions[get_column_letter(j)].width = width

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nombre_archivo() -> str:
    return f"sugerido_{date.today():%Y%m%d}.xlsx"
