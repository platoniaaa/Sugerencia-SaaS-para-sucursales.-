"""Lectores de los reportes de Flexline exportados a Excel (los que viven en SharePoint).

Puente entre los archivos que el usuario deja en SharePoint y el esquema CRUDO que
esperan las funciones de `conectores.sql_flexline` (que ya replican, y tienen
testeadas, las transformaciones del modelo). Con estos lectores el motor **no
necesita el SQL de Flexline**: alcanza con los Excel.

    Excel de SharePoint  ->  [este modulo]  ->  crudo con nombres del conector
                         ->  normalizar_seguimiento_* / normalizar_ventas_*  ->  motor

Particularidades de estos reportes (verificadas contra los archivos reales del
20-jul-2026):

- La fila de encabezados **no esta fija**: el reporte trae un titulo y filtros
  arriba (importado: fila 9; frontera: fila 8; ventas: fila 0). Por eso se
  DETECTA buscando la primera fila que contenga las columnas obligatorias, en vez
  de hardcodear un numero que se rompe al primer cambio de plantilla.
- La columna A viene vacia (los datos empiezan en la B).
- Las fechas de los seguimientos son texto `dd/mm/aaaa`; las de ventas son
  datetime real. Se aceptan ambos.
"""
from __future__ import annotations

import datetime as dt
import re
import unicodedata
from collections.abc import Iterable, Mapping
from pathlib import Path

import polars as pl

# Filas que se escanean buscando los encabezados antes de darse por vencido.
MAX_FILAS_ESCANEO = 40


def _norm(s: object) -> str:
    """Nombre de columna comparable: sin tildes, sin simbolos, minusculas.

    Asi `N° Orden de Compra`, `N Orden de Compra` y `n° orden de compra` son lo
    mismo, que es como cambian estos reportes entre exportaciones."""
    if s is None:
        return ""
    txt = unicodedata.normalize("NFKD", str(s).strip().lower())
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return "".join(c for c in txt if c.isalnum())


def _indice_de(origen: object, exactos: dict[str, int], normalizados: dict[str, int]) -> int | None:
    """Indice de la columna `origen`, prefiriendo la coincidencia LITERAL.

    Hay reportes con dos columnas distintas que normalizan igual: el respaldo de
    ventas trae `tipoproducto` (REPUESTO / REPUESTOS / MO_ST) y `Tipo Producto`
    (CAMION, RUBRO 70, SERVICIO TECNICO), y no significan lo mismo. Resolviendo
    solo por nombre normalizado ganaba la ultima y el filtro de repuestos se
    aplicaba sobre la columna equivocada: un aceite con 1.158 ventas quedaba
    fuera del sugerido porque su `Tipo Producto` decia CAMION.
    """
    j = exactos.get(str(origen).strip())
    return j if j is not None else normalizados.get(_norm(origen))


def _abrir_hoja(ruta: str | Path, hoja: str | None = None):
    import openpyxl  # import perezoso: solo se necesita al leer Excel

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    if hoja is not None and hoja in wb.sheetnames:
        return wb, wb[hoja]
    return wb, wb[wb.sheetnames[0]]


def leer_reporte(
    ruta: str | Path,
    columnas: Mapping[str, str],
    *,
    hoja: str | None = None,
    obligatorias: Iterable[str] | None = None,
) -> pl.DataFrame:
    """Lee un reporte de Flexline y devuelve solo las columnas pedidas, renombradas.

    `columnas` mapea nombre_destino -> nombre tal como aparece en el Excel. Las que
    no esten en el archivo salen como nulo (los tres seguimientos no traen las
    mismas columnas y el esquema del motor tiene que cuadrar igual).

    `obligatorias` son los nombres DESTINO que deben existir si o si; se usan para
    reconocer la fila de encabezados y para fallar temprano con un mensaje claro
    si el reporte cambio de formato. Por defecto, todas las de `columnas`.
    """
    ruta = Path(ruta)
    requeridas = [columnas[d] for d in (obligatorias if obligatorias is not None else columnas)]
    wb, ws = _abrir_hoja(ruta, hoja)
    try:
        filas = ws.iter_rows(values_only=True)
        indices: dict[str, int] | None = None
        for i, fila in enumerate(filas):
            if i >= MAX_FILAS_ESCANEO:
                break
            # Dos mapas: por nombre literal y por nombre normalizado. `setdefault`
            # para que gane la PRIMERA y no la ultima; el literal manda (ver
            # `_indice_de`: hay reportes con columnas que normalizan igual).
            exactos: dict[str, int] = {}
            normalizados: dict[str, int] = {}
            for j, v in enumerate(fila):
                if v is None:
                    continue
                exactos.setdefault(str(v).strip(), j)
                normalizados.setdefault(_norm(v), j)
            if all(_indice_de(o, exactos, normalizados) is not None for o in requeridas):
                indices = {
                    destino: j
                    for destino, origen in columnas.items()
                    if (j := _indice_de(origen, exactos, normalizados)) is not None
                }
                break
        if indices is None:
            raise ValueError(
                f"No se encontro la fila de encabezados en {ruta.name} "
                f"(se revisaron {MAX_FILAS_ESCANEO} filas). Faltan columnas como "
                f"{sorted(columnas[d] for d in (obligatorias or columnas))}."
            )

        datos: dict[str, list] = {destino: [] for destino in columnas}
        ancho_max = max(indices.values()) if indices else 0
        for fila in filas:  # el iterador sigue DESPUES del header
            if fila is None or len(fila) <= ancho_max:
                continue
            # Fila de relleno/subtotal: sin la primera columna obligatoria no hay dato.
            if all(fila[j] is None for j in indices.values()):
                continue
            for destino in columnas:
                j = indices.get(destino)
                datos[destino].append(None if j is None else fila[j])
    finally:
        wb.close()

    # Todo a texto: openpyxl mezcla int/str/datetime en una misma columna segun la
    # celda, y polars no puede inferir un tipo para eso. El casteo lo hace cada
    # lector segun lo que la columna significa.
    #
    # Con `.strip()`: los reportes traen codigos con espacios al final ("15 MXD1454M ")
    # y en el modelo eso NO crea un producto aparte -DAX compara texto ignorando el
    # relleno final-, pero en polars si. Sin recortar salian filas duplicadas por
    # producto-sucursal y esas copias no encontraban catalogo ni costo: 66 filas
    # sin Descripcion, Unidad de Medida ni FILTRO1_Final.
    return pl.DataFrame(
        {
            destino: [None if v is None else (v.isoformat() if isinstance(v, (dt.datetime, dt.date)) else str(v).strip())
                      for v in valores]
            for destino, valores in datos.items()
        },
        schema={destino: pl.Utf8 for destino in columnas},
    )


def _a_fecha(col: str) -> pl.Expr:
    """-> Date desde `dd/mm/aaaa`, ISO, o el serial numerico de Excel.

    Los respaldos anuales de ventas NO vienen todos igual: el de 2026 trae la
    fecha como datetime y el de 2025 como serial de Excel (45684 = 27-ene-2025).
    Sin el tercer camino el serial se leia como el texto "45684", no matcheaba
    ningun formato y el archivo entero quedaba con Fecha nula **en silencio**:
    198.032 ventas, medio ano de la ventana de demanda, perdidas sin un error.
    """
    c = pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
    serial = pl.col(col).cast(pl.Float64, strict=False)
    return pl.coalesce(
        c.str.to_date("%d/%m/%Y", strict=False),
        c.str.to_date("%Y-%m-%d", strict=False),
        c.str.head(10).str.to_date("%Y-%m-%d", strict=False),
        # Serial de Excel: el origen es 30-dic-1899 por el bug del 1900 bisiesto.
        # El rango acota a fechas plausibles (1954-2119) para no convertir por
        # accidente un numero que no era una fecha.
        pl.when(serial.is_between(20000, 80000))
        .then(pl.lit(dt.date(1899, 12, 30)) + pl.duration(days=serial.cast(pl.Int64)))
        .otherwise(None),
    ).alias(col)


def _a_entero(col: str) -> pl.Expr:
    return pl.col(col).cast(pl.Float64, strict=False).cast(pl.Int64).alias(col)


# --------------------------------------------------------------------------- #
# Seguimientos de compra
# --------------------------------------------------------------------------- #
# Nombre destino -> nombre en el Excel. Los destinos son los que espera
# `sql_flexline.normalizar_seguimiento*` (mismo esquema que trae el SQL).
COLUMNAS_SEGUIMIENTO_NACIONAL = {
    "Producto": "Producto",
    # El reporte trae TRES columnas de local: "Sucursal", "Código Local" y "Nombre
    # Local", y no son lo mismo. El modelo deriva SucursalID de "Sucursal"; usar
    # "Código Local" mandaba 6.964 ordenes a DESCONOCIDO (contra 1.093) y repartia
    # las de un mismo local entre varias sucursales, lo que descuadraba el lead
    # time por proveedor-sucursal.
    "Sucursal": "Sucursal",
    "RazonSocial": "Razón Social Proveedor",
    "Motivo": "Motivo Compra",
    "FechaOC": "Fecha Orden de Compra",
    "NOC": "N° Orden de Compra",
    "Cantidad": "Cantidad",
    "EstadoOC": "Estado Orden de Compra",
    "EstadoDoc": "Estado Documento Base",
    "FechaDoc": "Fecha Documento Base",
    "FechaPE": "Fecha Documento P/E",
}

COLUMNAS_SEGUIMIENTO_IMPORTADO = {
    "Producto": "Producto",
    # El importado trae [Sucursal] casi siempre en blanco -> DESCONOCIDO. Es el
    # comportamiento del modelo (solo aporta al fallback global de proveedor); NO
    # usar "Código Local", que aca viene con el NOMBRE y no con el codigo SUC0XX.
    "Sucursal": "Sucursal",
    "RazonSocial": "Razón Social Proveedor",
    "Motivo": "Motivo Compra",
    "FechaOC": "Fecha Orden de Compra",
    "NOC": "N° Orden de Compra",
    "Cantidad": "Cantidad",
    "EstadoDoc": "Estado Documento Base",
    "FechaDoc": "Fecha Documento Base",
    # El importado no tiene "Fecha Documento P/E": la recepcion de la importacion
    # es el equivalente (es la fecha en que la mercaderia queda disponible).
    "FechaPE": "Fecha Documento Recepción",
    "EstadoOC": "Estado Documento Recepción",
}

COLUMNAS_SEGUIMIENTO_FRONTERA = {
    "Producto": "Producto",
    "NombreLocal": "Nombre Local",
    "RazonSocial": "Razón Social Proveedor",
    "Cantidad": "Cantidad",
    "EstadoDoc": "Estado Documento Base",
    "FechaDoc": "Fecha Documento Base",
    # Frontera no trae OC como tal: el documento base ES la orden de compra.
    "FechaOC": "Fecha Documento Base",
    "NOC": "N° Documento Base",
    "FechaPE": "Fecha Recepción",
}

_FECHAS_SEGUIMIENTO = ("FechaOC", "FechaDoc", "FechaPE")


def _leer_seguimiento(ruta: str | Path, columnas: Mapping[str, str]) -> pl.DataFrame:
    df = leer_reporte(
        ruta, columnas, obligatorias=[d for d in ("Producto", "Cantidad") if d in columnas]
    )
    return df.with_columns(
        [_a_fecha(c) for c in _FECHAS_SEGUIMIENTO if c in df.columns]
        + [_a_entero("Cantidad")]
    )


def leer_seguimiento_nacional_excel(ruta: str | Path) -> pl.DataFrame:
    """'Seguimiento de Compras' de Curifor nacional (el que hoy sale del SQL)."""
    return _leer_seguimiento(ruta, COLUMNAS_SEGUIMIENTO_NACIONAL)


def leer_seguimiento_importado_excel(ruta: str | Path) -> pl.DataFrame:
    """'Seguimiento Compras' de importaciones (O/C IMPORTACION)."""
    return _leer_seguimiento(ruta, COLUMNAS_SEGUIMIENTO_IMPORTADO)


def leer_seguimiento_frontera_excel(ruta: str | Path) -> pl.DataFrame:
    """'Seguimiento de Compras' de Frontera."""
    return _leer_seguimiento(ruta, COLUMNAS_SEGUIMIENTO_FRONTERA)


# --------------------------------------------------------------------------- #
# Ventas (respaldos anuales de SharePoint)
# --------------------------------------------------------------------------- #
# Las columnas crudas que consume `normalizar_ventas_curifor`.
COLUMNAS_VENTAS = {
    "Producto": "Producto",
    "SUCURSAL": "SUCURSAL",
    "Tipo-Venta": "Tipo-Venta",
    "Fecha": "Fecha",
    "Cantidad": "Cantidad",
    "tipoDocto": "tipoDocto",
    "tipoproducto": "tipoproducto",
    "Empresa": "Empresa",
    # La descripcion del producto sale de las VENTAS, no del listado maestro
    # (DAX: 'Dim Producto'[Descripcion] = MINX sobre 'Ventas Unificadas').
    "Descripcion Producto": "Descripcion Producto",
}

# Valor canonico que espera el filtro de repuestos del conector.
_TIPOPRODUCTO_REPUESTOS = "REPUESTOS"


def _tipoproducto_canonico() -> pl.Expr:
    """`4Repuesto` (respaldo Excel) -> `REPUESTOS` (lo que devuelve el SQL).

    El respaldo anual clasifica con un prefijo de orden (`1M.O.`, `3Insumo`,
    `4Repuesto`, `5 Adicional`) que el SQL no tiene. Sin esta equivalencia el
    filtro de repuestos del conector no matchea NADA y las ventas salen vacias."""
    return (
        pl.when(pl.col("tipoproducto").str.to_lowercase().str.contains("repuesto"))
        .then(pl.lit(_TIPOPRODUCTO_REPUESTOS))
        .otherwise(pl.col("tipoproducto"))
        .alias("tipoproducto")
    )


def _sucursal_sin_prefijo() -> pl.Expr:
    """`08 TALCA` -> `TALCA`. El respaldo trae la sucursal con el prefijo de orden
    del informe; el modelo la resuelve via Dim_Locales y el sugerido usa el nombre
    pelado. Los valores sin prefijo (`DIEZ DE JULIO (2)`) quedan igual."""
    return (
        pl.col("SUCURSAL")
        .str.replace(r"^\d{1,2}\s+", "")
        .alias("SUCURSAL")
    )


def leer_ventas_excel(rutas: str | Path | Iterable[str | Path]) -> pl.DataFrame:
    """Respaldos anuales de ventas -> crudo de `normalizar_ventas_curifor`.

    Acepta una ruta o varias (los respaldos vienen partidos por ano: `2024.xlsx`,
    `2020 2023.xlsx`, ...) y las concatena. Ademas de tipar Fecha y Cantidad,
    traduce al vocabulario del SQL las dos columnas donde el respaldo Excel usa
    otro formato (`tipoproducto` y `SUCURSAL`), para que las reglas del modelo
    ya implementadas en `sql_flexline` apliquen sin cambios.
    """
    if isinstance(rutas, (str, Path)):
        rutas = [rutas]
    frames = []
    for ruta in rutas:
        df = leer_reporte(
            ruta, COLUMNAS_VENTAS, obligatorias=["Producto", "Cantidad", "tipoDocto", "tipoproducto"]
        )
        frames.append(
            df.with_columns(
                _a_fecha("Fecha"),
                _a_entero("Cantidad"),
                _tipoproducto_canonico(),
                _sucursal_sin_prefijo(),
            )
        )
    if not frames:
        raise ValueError("leer_ventas_excel necesita al menos un archivo")
    return frames[0] if len(frames) == 1 else pl.concat(frames, how="vertical_relaxed")


# --------------------------------------------------------------------------- #
# Ventas Frontera (E07)
# --------------------------------------------------------------------------- #
# Los nombres son los del SELECT que el modelo hace contra Flexline, asi que el
# reporte exportado los trae tal cual. OJO: igual que el respaldo Curifor, trae
# DOS columnas de tipo de producto -`tipoproducto` (REPUESTO / MO_ST) y
# `TIPO PRODUCTO` (REPUESTOS / MECANICA / LUBRICANTE)- que normalizan al mismo
# nombre. El modelo filtra por la primera; `leer_reporte` la resuelve por
# coincidencia literal (ver `_indice_de`).
COLUMNAS_VENTAS_FRONTERA = {
    "producto": "producto",
    "SUCURSAL": "SUCURSAL",
    "Tipo-Venta": "Tipo-Venta",
    "fecha": "fecha",
    "cantidad": "cantidad",
    "Documento": "Documento",
    "Docto-Emitido": "Docto-Emitido",
    "tipoproducto": "tipoproducto",
}


def leer_ventas_frontera_excel(ruta: str | Path) -> pl.DataFrame:
    """'Informe Gestion Produccion REP ST GAR D&P' (E07) -> crudo de
    `normalizar_ventas_frontera`, que aplica los filtros del modelo."""
    df = leer_reporte(
        ruta,
        COLUMNAS_VENTAS_FRONTERA,
        obligatorias=["producto", "cantidad", "Documento", "Docto-Emitido", "tipoproducto"],
    )
    return df.with_columns(_a_fecha("fecha"), _a_entero("cantidad"))


# --------------------------------------------------------------------------- #
# Listado maestro de repuestos y base de reemplazos
# --------------------------------------------------------------------------- #
# Columnas que el motor usa del maestro. Las dos primeras son obligatorias.
COLS_MAESTRO = ("Producto", "Categoria", "Glosa", "Unidad", "Familia")

# Hojas donde ha vivido el maestro dentro de la planilla de precios, en orden de
# preferencia. Se comparan normalizadas (sin tildes ni simbolos).
HOJAS_LISTADO_MAESTRO = ("Lista sin duplicados", "Listado Maestro", "Maestro")


def _hoja_maestro(disponibles: list[str], hoja: str | None) -> str | None:
    """Nombre REAL de la hoja del maestro dentro del libro, o None si no esta."""
    por_norm = {_norm(n): n for n in disponibles}
    for candidata in ([hoja] if hoja else HOJAS_LISTADO_MAESTRO):
        real = por_norm.get(_norm(candidata))
        if real is not None:
            return real
    return None


def _leer_maestro_excel(ruta: Path, hoja: str | None) -> pl.DataFrame:
    import fastexcel  # el motor que usa pl.read_excel; aca se usa para elegir la hoja

    libro = fastexcel.read_excel(str(ruta))
    elegida = _hoja_maestro(libro.sheet_names, hoja)
    if elegida is None:
        raise ValueError(
            f"{ruta.name} no trae la hoja del listado maestro (se buscaron "
            f"{list(HOJAS_LISTADO_MAESTRO)}). Hojas del archivo: {libro.sheet_names}. "
            "Si el archivo no es el maestro, revisa los patrones de la fuente 'catalogo'."
        )
    # Se leen SOLO las columnas que el motor usa: la hoja real trae 52 columnas y
    # 410.000 filas, y pedir las cinco que importan baja la lectura de minutos a
    # segundos. El sondeo previo (n_rows=0) es para tolerar que falte una opcional.
    presentes = {c.name for c in libro.load_sheet(elegida, n_rows=0).available_columns()}
    return pl.read_excel(
        ruta, sheet_name=elegida, columns=[c for c in COLS_MAESTRO if c in presentes]
    )


def leer_listado_maestro(ruta: str | Path, hoja: str | None = None) -> pl.DataFrame:
    """'Listado Maestro Repuestos' -> Producto, Categoria (+ Glosa, Unidad, Familia).

    De aqui sale la Categoria que deja COLISION y CAMPANAS fuera del sugerido. La
    fuente ha venido de dos formas y se aceptan las dos:

    - CSV (export de Flexline, "lista rep (lista precios).csv"): viene con `;`, en
      latin-1 y SIN entrecomillar; hay glosas con comillas sueltas (medidas en
      pulgadas) que rompen el parser si se interpretan como delimitador de texto,
      por eso `quote_char=None`.
    - Excel ("LISTA DE PRECIOS.xlsx", hoja "Lista sin duplicados"): desde jul-2026
      el maestro dejo de exportarse aparte y vive como una hoja de la planilla de
      precios de Curifor.

    Todo sale como texto, como salia del CSV con `infer_schema_length=0`: el Excel
    infiere tipos por columna y aguas abajo se compara y se joinea contra texto.
    """
    ruta = Path(ruta)
    if ruta.suffix.lower() == ".csv":
        df = pl.read_csv(
            ruta, separator=";", encoding="latin-1", quote_char=None, infer_schema_length=0
        )
    else:
        df = _leer_maestro_excel(ruta, hoja)
    faltan = {"Producto", "Categoria"} - set(df.columns)
    if faltan:
        raise ValueError(
            f"El listado maestro {ruta.name} no trae {sorted(faltan)}. "
            f"Columnas encontradas: {df.columns[:12]}"
        )
    cols = [c for c in COLS_MAESTRO if c in df.columns]
    return df.select(cols).with_columns(
        pl.all().cast(pl.Utf8, strict=False)
    ).with_columns(pl.col("Producto").str.strip_chars())


# La Hoja2 del "BASE NUEVO MIX" es la que lee el modelo (no Hoja1, que es una tabla
# dinamica, ni BBDD). Trae una fila en blanco antes del encabezado.
COLUMNAS_MIX = {
    "Producto": "Producto",
    "Reem1": "Reem1",
    "Reem2": "Reem2",
    "Reem3": "Reem3",
}


def leer_mix_reemplazos(ruta: str | Path) -> pl.DataFrame:
    """'mix andres' -> Producto, Reem1, Reem2, Reem3 (base de los grupos de reemplazo)."""
    return leer_reporte(ruta, COLUMNAS_MIX, hoja="Hoja2", obligatorias=["Producto", "Reem1"])


# --- Listas de precios de proveedor -------------------------------------------
# No alimentan el sugerido: dan el precio de VENTA para calcular el margen y poder
# priorizar que comprar. Se cruzan con el codigo interno de Curifor, que es
# "<prefijo numerico> <codigo del fabricante>" ("25 DG9Z8100A"), mientras que las
# listas traen el codigo del fabricante en su propio formato ("DG9Z/8100/A/").

def clave_precio(codigo: str | None) -> str | None:
    """Codigo comparable entre las listas y el maestro de Curifor.

    Saca el prefijo numerico de Curifor y deja solo letras y numeros en mayuscula,
    con lo que "25 DG9Z8100A" y "DG9Z/8100/A/" caen en la misma clave. Sin esto el
    cruce da 0: FORD separa con barras y Curifor no.
    """
    if codigo is None:
        return None
    s = str(codigo).strip()
    if not s:
        return None
    m = re.match(r"^\d+\s+(.*)$", s)
    if m:
        s = m.group(1)
    limpio = re.sub(r"[^A-Za-z0-9]", "", s).upper()
    return limpio or None


# Columna de la lista -> columna del contrato de la plataforma.
COLUMNAS_PRECIOS_FORD = {
    "Price_dealer": "precio_dealer_ford",
    "Precio_Publico": "precio_publico_ford",
    "Precio_Publico_ConImpuestos": "precio_publico_iva_ford",
    "Reposicion": "precio_reposicion_ford",
    "Urgente_VOR": "precio_urgente_vor_ford",
    "Promociones": "precio_promociones_ford",
    "Urgente_Recargo15": "precio_urgente_recargo15_ford",
    "Precio_Flota": "precio_flota_ford",
}

COLUMNAS_PRECIOS_GILDEMEISTER = {
    "Precio_Sugerido": "precio_sugerido_gilde",
    "Precio_Dealer": "precio_dealer_gilde",
    "Precio_Final_Dealer": "precio_final_dealer_gilde",
}


def _leer_precios(ruta: str | Path, col_codigo: str, mapa: dict[str, str]) -> pl.DataFrame:
    """Hoja 'Precios' de una lista de proveedor -> clave + columnas de precio."""
    df = pl.read_excel(ruta, sheet_name="Precios")
    if col_codigo not in df.columns:
        raise ValueError(
            f"La lista de precios {Path(ruta).name} no trae la columna {col_codigo!r}. "
            f"Columnas encontradas: {df.columns[:12]}"
        )
    presentes = {k: v for k, v in mapa.items() if k in df.columns}
    if not presentes:
        raise ValueError(
            f"La lista de precios {Path(ruta).name} no trae ninguna columna de precio "
            f"conocida ({sorted(mapa)}). Columnas encontradas: {df.columns[:12]}"
        )
    out = df.select(
        pl.col(col_codigo).alias("_codigo"),
        *[pl.col(k).cast(pl.Float64, strict=False).alias(v) for k, v in presentes.items()],
    ).with_columns(
        pl.col("_codigo")
        .map_elements(clave_precio, return_dtype=pl.Utf8)
        .alias("clave_precio")
    ).drop("_codigo")
    # Una clave puede repetirse si dos codigos distintos se normalizan igual; con
    # el precio mas alto se evita subestimar el margen (y es estable entre corridas).
    return (
        out.filter(pl.col("clave_precio").is_not_null())
        .group_by("clave_precio")
        .agg([pl.col(c).max() for c in presentes.values()])
    )


def leer_precios_ford(ruta: str | Path) -> pl.DataFrame:
    return _leer_precios(ruta, "PartNumber", COLUMNAS_PRECIOS_FORD)


# Columnas de reemplazo de la extraccion FORD WINGS (desde ago-2026). Vienen en la
# misma hoja 'Precios' y son opcionales: una lista vieja sin ellas se lee igual.
COLUMNAS_REEMPLAZO_FORD = [
    "Reemplazado_Por", "Cadena_Reemplazo", "Reemplaza_A",
    "Estado_Reemplazo", "Reemplazo_Aviso", "Fecha_Extraccion",
]

# Forma con la que el motor consume los reemplazos de FORD, venga de donde venga.
# La lista de precios y la consulta en vivo de WINGS tienen que devolver EXACTAMENTE
# esto: `combinar_reemplazos_ford` las mezcla campo a campo y
# `dimensiones.ampliar_mapeo_con_ford` consume el resultado sin saber de cual de las
# dos salio cada dato. Si una de las dos se desalinea, la mezcla falla en silencio.
ESQUEMA_REEMPLAZOS_FORD: dict[str, pl.DataType] = {
    "clave_precio": pl.Utf8, "sku_ford": pl.Utf8, "clave_vigente": pl.Utf8,
    "sku_vigente": pl.Utf8, "cadena": pl.Utf8, "reemplaza_a": pl.List(pl.Utf8),
    "estado_reemplazo": pl.Utf8, "sucesor_confirmado": pl.Boolean,
    "aviso": pl.Utf8, "extraido_en": pl.Utf8,
}


def leer_reemplazos_ford(ruta: str | Path) -> pl.DataFrame:
    """Cadena de reemplazo que publica FORD, en claves comparables con Curifor.

    Devuelve una fila por codigo de la lista con:
      clave_precio, sku_ford, clave_vigente, sku_vigente, cadena, reemplaza_a
      (lista de claves), estado_reemplazo, sucesor_confirmado, aviso.

    Dos direcciones, porque la lista trae las dos y no se solapan:
      - `Reemplazado_Por`: a ESTA pieza la reemplaza otra (1.070 codigos).
      - `Reemplaza_A`: esta pieza reemplazo a otras (11.025 codigos), que es de
        donde sale casi todo lo que Curifor efectivamente tiene.

    `sucesor_confirmado` es False cuando FORD dice "Sin candidato vigente" (999 de
    los 1.070). Gobierna DOS cosas y por eso no se llama "precio_...": el precio
    del sucesor no viene, y ademas el codigo del sucesor no esta verificado -- no
    se sabe si FORD realmente no lo tiene o si el codigo consultado se armo mal.
    Sin esa confirmacion no se puede agrupar: juntar el stock de dos piezas que no
    son la misma es peor que no juntarlo, porque deja de pedirse algo que si hace
    falta. `Estado_Reemplazo` aplica solo a `Reemplazado_Por`; `Reemplaza_A` viene
    de la cadena del portal y no depende de esa consulta.

    Si la hoja no trae las columnas, devuelve un frame vacio con el mismo esquema.
    """
    df = pl.read_excel(ruta, sheet_name="Precios")
    if not any(c in df.columns for c in COLUMNAS_REEMPLAZO_FORD):
        return pl.DataFrame(schema=ESQUEMA_REEMPLAZOS_FORD)

    def _txt(col: str) -> pl.Expr:
        if col not in df.columns:
            return pl.lit(None, dtype=pl.Utf8).alias(col)
        return pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars().replace("", None)

    out = df.select(
        pl.col("PartNumber").cast(pl.Utf8).alias("sku_ford"),
        *[_txt(c) for c in COLUMNAS_REEMPLAZO_FORD],
    ).with_columns(
        pl.col("sku_ford").map_elements(clave_precio, return_dtype=pl.Utf8).alias("clave_precio"),
        pl.col("Reemplazado_Por").alias("sku_vigente"),
        pl.col("Reemplazado_Por")
        .map_elements(clave_precio, return_dtype=pl.Utf8)
        .alias("clave_vigente"),
        pl.col("Cadena_Reemplazo").alias("cadena"),
        pl.col("Estado_Reemplazo").alias("estado_reemplazo"),
        pl.col("Reemplazo_Aviso").alias("aviso"),
        # Cuando se consulto el portal por ESTA fila. Va por fila y no como un
        # valor global porque el motor combina dos archivos -la lista de FORD y
        # la de los codigos de Curifor- y cada uno se extrae por su lado: una
        # fila puede ser de hoy y la de al lado de hace tres semanas.
        pl.col("Fecha_Extraccion").alias("extraido_en"),
        # Solo "Encontrado" deja el sucesor confirmado (codigo y precio).
        (pl.col("Estado_Reemplazo") == "Encontrado").alias("sucesor_confirmado"),
        # "A; B; C" -> claves normalizadas, sin vacios.
        pl.col("Reemplaza_A")
        .str.split(";")
        .list.eval(
            pl.element().str.strip_chars().map_elements(clave_precio, return_dtype=pl.Utf8)
        )
        .list.drop_nulls()
        .alias("reemplaza_a"),
    )
    return out.select(list(ESQUEMA_REEMPLAZOS_FORD)).filter(
        pl.col("clave_precio").is_not_null()
    )


def _colapsar_por_clave(df: pl.DataFrame) -> pl.DataFrame:
    """Una fila por `clave_precio`, la mas informativa.

    El mismo repuesto puede venir dos veces partido distinto -`8A61/A03195AE5/YY/`
    y `8A61/A03195/AE/5YY` son el mismo numero de parte con el slash en otro lado-
    y las dos formas dan la misma `clave_precio`. Al 22-08-2026: 6 casos en la
    lista de FORD y 60 en la corrida de WINGS, donde el traductor consulta las dos
    particiones cuando no esta seguro.

    Sin colapsar pasan dos cosas, las dos malas:

      - `combinar_reemplazos_ford` hace un `left join` por esta clave, y una clave
        repetida del lado derecho MULTIPLICA la fila del izquierdo;
      - la plataforma no tiene clave unica por producto, asi que dos filas del
        mismo codigo la dejan quedandose con cualquiera. En `AB3917D698AC3ZH` una
        de las dos trae el vigente y la otra no: el aviso al comprador aparecia o
        no segun el orden de insercion.

    Gana la fila que resolvio sucesor; despues la que nombra mas predecesores;
    despues la mas reciente. El ultimo criterio es el `sku_ford` para que dos
    corridas iguales elijan igual.
    """
    if df.height == df["clave_precio"].n_unique():
        return df
    return (
        df.sort(
            by=[
                pl.col("sku_vigente").is_not_null(),
                pl.col("reemplaza_a").list.len().fill_null(0),
                pl.col("extraido_en"),
                pl.col("sku_ford"),
            ],
            descending=[True, True, True, False],
            nulls_last=True,
        )
        .unique(subset="clave_precio", keep="first", maintain_order=True)
    )


def combinar_reemplazos_ford(lista: pl.DataFrame, wings: pl.DataFrame) -> pl.DataFrame:
    """Mezcla la lista de precios estatica con lo que WINGS consulto en vivo.

    Cada fuente manda en una direccion distinta y por eso no se pisa la fila entera:

    - **WINGS manda hacia adelante** (`clave_vigente`: a este codigo lo reemplaza X).
      Consulta el portal en el momento, asi que resuelve los descontinuados que la
      foto no tiene. Si WINGS dice que un codigo NO tiene reemplazo, eso tambien
      manda: borra el sucesor que la lista de precios declaraba.
    - **La lista de precios manda en la inversa** (`reemplaza_a`: este codigo
      reemplazo a A, B, C), que WINGS no publica y de donde sale casi todo lo que
      Curifor efectivamente agrupa.

    Pisar la fila completa se veria mas simple y estaria mal: de los 33 codigos de
    la pauta InStock, 11 traen `reemplaza_a` en la lista de precios (medido el
    22-08-2026). Reemplazarlos enteros perderia esos 11 grupos sin que nada avisara.
    """
    # Antes de cruzar nada: las dos fuentes traen el mismo repuesto repetido con
    # otra particion, y una clave repetida rompe el join de abajo. Ver
    # `_colapsar_por_clave`.
    lista = _colapsar_por_clave(lista)
    wings = _colapsar_por_clave(wings)

    if wings.is_empty():
        return lista
    cols = list(ESQUEMA_REEMPLAZOS_FORD)
    if lista.is_empty():
        return wings.select(cols)

    # Las filas que WINGS no toca quedan intactas.
    solo_lista = lista.join(wings.select("clave_precio"), on="clave_precio", how="anti")
    # Las que si toca: todo de WINGS, menos `reemplaza_a`, que se trae de la lista.
    # `coalesce` sirve porque una lista vacia no es null: para un codigo que WINGS
    # conoce y la lista no, gana el [] de WINGS, que es lo correcto.
    de_wings = (
        wings.join(
            lista.select(["clave_precio", pl.col("reemplaza_a").alias("_ra_lista")]),
            on="clave_precio",
            how="left",
        )
        .with_columns(pl.coalesce("_ra_lista", "reemplaza_a").alias("reemplaza_a"))
        .select(cols)
    )
    return pl.concat([solo_lista.select(cols), de_wings])


def leer_precios_gildemeister(ruta: str | Path) -> pl.DataFrame:
    return _leer_precios(ruta, "Codigo", COLUMNAS_PRECIOS_GILDEMEISTER)
