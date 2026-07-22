"""Orquestador de FUENTES REALES: arma el dict `fuentes` de `pipeline.ejecutar`
leyendo de los crudos reales en vez de los CSV extraídos del modelo.

Cada entrada del motor y de dónde sale su crudo real (ver FUENTES_REALES.md):

| entrada del motor          | fuente real                              | estado |
|----------------------------|------------------------------------------|--------|
| stock / stock_frontera     | Stock bodegas[ frontera].xlsx (Excel)    | ✅ `leer_stock` |
| costo                      | mismo Excel de stock (columna Costo)     | ✅ `leer_stock` |
| seguimiento / _lt / _transito | Excel de SharePoint (nacional/importado/frontera) | ✅ `lectores_excel` |
| ventas                     | respaldos anuales de SharePoint (Excel)  | ✅ `lectores_excel` |
| mapeo_master               | 'mix andres' (BASE NUEVO MIX.xlsx) → DAX complejo | ⏳ snapshot o replicar |
| dim_producto               | tabla calculada del modelo (ventas+stock+catálogo) | ⏳ snapshot o replicar |
| dim_sucursal / importados  | tablas chicas y estables                 | ⏳ snapshot |

El SQL de Flexline (`conectores/sql_flexline.py`) queda como camino alternativo: sus
transformaciones se siguen usando (son las reglas del modelo), pero las fuentes de
alta frecuencia entran por Excel, que es lo que el usuario puede publicar a diario
en SharePoint sin depender de credenciales del ERP ni de estar en la LAN.

Las tablas chicas y estables (mapeo, dim_producto, dim_sucursal, importados) se
pueden **snapshotear** (leer de un CSV congelado que se regenera de vez en cuando)
sin perder frescura, porque cambian poco. Lo que SÍ necesita conexión viva es lo
de alta frecuencia: ventas, stock y seguimiento.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import polars as pl

# --- Bodega -> SucursalID (SWITCH de 'Stock Bodegas'[SucursalID], case-insensitive) ---
STOCK_BODEGA_SUCURSAL = {
    "AUTOPARK": "AUTOPARK", "BDGA. GRAN AVENIDA": "GRAN AVENIDA",
    "BODEGA DAÑADOS": "BODEGA DANADOS", "BODEGA DAÃ'ADOS": "BODEGA DANADOS",
    "BODEGA DEVOLUCION": "BODEGA DEVOLUCION", "BODEGA DYP CHILLAN V": "CHILLAN VIEJO",
    "BODEGA DYP CURICO": "CURICO", "BODEGA DYP PLACILLA": "PLACILLA",
    "BODEGA DYP TALCA": "TALCA", "BODEGA DYP TALCA 2": "TALCA (2)",
    "BODEGA IMPORTACION": "BODEGA IMPORTACION", "BODEGA ML": "CD REPUESTOS",
    "BODEGA ML-FULL": "CD REPUESTOS", "BODEGA SCRAP": "BODEGA SCRAP",
    "BRASIL 18": "BRASIL 18", "CD REPUESTOS": "CD REPUESTOS", "CHILLAN": "CHILLAN",
    "CHILLAN2": "CHILLAN VIEJO", "COMEX": "COMEX", "CURICO": "CURICO",
    "DIEZ DE JULIO": "DIEZ DE JULIO", "DIEZ DE JULIO (2)": "DIEZ DE JULIO (2)",
    "IMPORTACION MOTOS": "IMPORTACION MOTOS", "LA FLORIDA": "LA FLORIDA",
    "LINDEROS": "LINDEROS", "LO BLANCO": "LO BLANCO", "LO BLANCO 2": "LO BLANCO",
    "MALL PLAZA NORTE": "MALL PLAZA NORTE", "MALL PLAZA SUR": "MALL PLAZA SUR",
    "OVALLE": "OVALLE", "PE X REGULARIZAR": "PE X REGULARIZAR", "PE-FALTANTE": "PE FALTANTE",
    "PLACILLA": "PLACILLA", "RANCAGUA": "RANCAGUA", "RANCAGUA 2": "RANCAGUA",
    "RANCAGUA 3": "RANCAGUA", "ST_RANCAGUA": "RANCAGUA", "TALCA": "TALCA",
    "TALCA (2)": "TALCA (2)", "TALCA BMW": "TALCA", "TRANSITO": "TRANSITO",
}
_BODEGA_LOWER = {k.lower(): v for k, v in STOCK_BODEGA_SUCURSAL.items()}
STOCK_SUCURSAL_DEFAULT = "DESCONOCIDO"


def leer_stock(ruta_excel: str | Path) -> pl.DataFrame:
    """Lee un 'Stock bodegas[ frontera].xlsx' (Hoja1) y devuelve un frame con
    Producto, SucursalID (mapeado desde Bodega), Stock y Costo — el mismo esquema
    (nivel de fila) que la tabla 'Stock Bodegas' del modelo. Réplica de su M:
    promover encabezados + SucursalID = SWITCH(Bodega) case-insensitive.

    De aquí salen a la vez `stock`/`stock_frontera` (Producto, SucursalID, Stock)
    y `costo` (Producto, Costo)."""
    import openpyxl  # import perezoso: solo se necesita al leer Excel

    wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)
    try:
        ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(it)]
        idx = {
            name: header.index(name)
            for name in ("Producto", "Bodega", "Stock", "Costo", "Familia")
            if name in header
        }
        faltan = {"Producto", "Bodega", "Stock"} - set(idx)
        if faltan:
            raise ValueError(f"Al Excel de stock le faltan columnas: {faltan}")
        prods, bods, stocks, costos, familias = [], [], [], [], []
        ic, ifam = idx.get("Costo"), idx.get("Familia")
        for row in it:
            if row is None or row[idx["Producto"]] is None:
                continue
            prods.append(row[idx["Producto"]])
            bods.append(row[idx["Bodega"]])
            stocks.append(row[idx["Stock"]])
            costos.append(row[ic] if ic is not None else None)
            # Familia trae "RUBRO N": de ahi sale el rubro que define FILTRO1_Final.
            familias.append(row[ifam] if ifam is not None else None)
    finally:
        wb.close()

    # Todo a texto al construir (openpyxl devuelve int/str mezclados por columna);
    # Stock se castea a entero y Costo queda texto (el modelo lo pasa por VALUE()).
    a_texto = lambda xs: [None if v is None else str(v) for v in xs]
    df = pl.DataFrame({
        "Producto": a_texto(prods), "Bodega": a_texto(bods),
        "Stock": a_texto(stocks), "Costo": a_texto(costos),
        "Familia": a_texto(familias),
    })
    return df.with_columns(
        pl.col("Bodega").fill_null("").str.to_lowercase()
        .replace_strict(_BODEGA_LOWER, default=STOCK_SUCURSAL_DEFAULT)
        .alias("SucursalID")
    ).select(["Producto", "SucursalID", pl.col("Stock").cast(pl.Float64, strict=False).cast(pl.Int64), "Costo"])


def _seguimiento_desde_excel(
    nacional_xlsx: str | Path | None,
    importado_xlsx: str | Path | None,
    frontera_xlsx: str | Path | None,
) -> dict[str, pl.DataFrame]:
    """Las tres vistas del seguimiento (base, transito y lead time) desde los Excel.

    Cada archivo se lee UNA vez y se normaliza tres veces, porque las tres vistas
    salen del mismo crudo con distintas columnas."""
    from . import lectores_excel as lx
    from .conectores import sql_flexline as sf

    crudos: list[tuple[pl.DataFrame, callable]] = []
    if nacional_xlsx is not None:
        crudos.append((lx.leer_seguimiento_nacional_excel(nacional_xlsx), sf.normalizar_seguimiento))
    if importado_xlsx is not None:
        crudos.append(
            (lx.leer_seguimiento_importado_excel(importado_xlsx), sf.normalizar_seguimiento_importado)
        )
    if frontera_xlsx is not None:
        crudos.append(
            (lx.leer_seguimiento_frontera_excel(frontera_xlsx), sf.normalizar_seguimiento_frontera)
        )
    if not crudos:
        raise ValueError("Se pidio el seguimiento desde Excel pero no se paso ningun archivo")

    def _union(**kwargs) -> pl.DataFrame:
        return sf.unir_seguimiento(*[normalizar(crudo, **kwargs) for crudo, normalizar in crudos])

    return {
        "seguimiento": _union(),
        "seguimiento_transito": _union(para_transito=True),
        "seguimiento_lt": _union(para_lead_time=True),
    }


def cargar_fuentes_reales(
    *,
    stock_curifor_xlsx: str | Path,
    stock_frontera_xlsx: str | Path,
    snapshot_dir: str | Path,
    seguimiento_nacional_xlsx: str | Path | None = None,
    seguimiento_importado_xlsx: str | Path | None = None,
    seguimiento_frontera_xlsx: str | Path | None = None,
    ventas_xlsx: str | Path | list[str | Path] | None = None,
    ventas_frontera_crudo: pl.DataFrame | None = None,
    listado_maestro: str | Path | None = None,
    mix_reemplazos_xlsx: str | Path | None = None,
    fin_mes_cerrado: date | None = None,
    sql_conn=None,
) -> dict[str, pl.DataFrame]:
    """Arma el dict `fuentes` para `pipeline.ejecutar` desde los crudos reales.

    - `stock_*_xlsx`: Excel de stock (Curifor y Frontera).
    - `snapshot_dir`: carpeta con los CSV congelados de las tablas chicas y estables
      (mapeo_master, dim_producto, dim_sucursal, importados).
    - `seguimiento_*_xlsx` / `ventas_xlsx`: Excel de SharePoint. **Es el camino
      principal**: con ellos el motor no necesita el SQL de Flexline ni la red de
      Curifor. Las ventas aceptan varios archivos (respaldos por ano).
    - `listado_maestro` / `mix_reemplazos_xlsx`: con estos dos, mas las ventas y el
      seguimiento importado, el motor CALCULA las tablas chicas en vez de leerlas
      del snapshot congelado, y deja de depender del Power BI. `fin_mes_cerrado`
      hace falta para el mapeo (elige el master del grupo por venta de 6 meses).
    - `sql_conn`: conexión a Flexline (`conectores.sql_flexline.conectar()`), como
      alternativa si algun dia se quiere leer en vivo desde la LAN.

    Para cada fuente se toma, en orden: el Excel si se paso, si no el SQL si hay
    conexión, si no el snapshot congelado. Asi se puede migrar de a una fuente sin
    quedarse a medias.
    """
    from . import dimensiones as dim

    snap = Path(snapshot_dir)
    S = {"Producto": pl.Utf8, "SucursalID": pl.Utf8}
    stock = leer_stock(stock_curifor_xlsx)
    stock_frontera = leer_stock(stock_frontera_xlsx)

    def _snapshot(nombre: str, **kw) -> pl.DataFrame | None:
        """Snapshot congelado del BI, si todavía está. Es solo un respaldo: las
        tablas chicas se calculan más abajo desde las fuentes propias, y el motor
        tiene que poder correr con esta carpeta vacía."""
        ruta = snap / nombre
        return pl.read_csv(ruta, **kw) if ruta.exists() else None

    fuentes: dict[str, pl.DataFrame] = {
        "stock": stock.select(["Producto", "SucursalID", "Stock"]),
        "stock_frontera": stock_frontera.select(["Producto", "SucursalID", "Stock"]),
        # Costo: del Excel de stock Curifor (columna Costo, como el modelo).
        "costo": stock.select(["Producto", pl.col("Costo").cast(pl.Utf8)]),
        # 'Dim Sucursal' es una DATATABLE escrita a mano dentro del modelo: no viene
        # de ninguna fuente, asi que vive en el codigo y no necesita snapshot.
        "dim_sucursal": dim.dim_sucursal(),
    }
    for nombre, archivo, kw in (
        ("mapeo", "mapeo_master.csv", {}),
        ("dim_producto", "dim_producto.csv", {"schema_overrides": S}),
        ("catalogo", "dim_producto_catalogo.csv", {"schema_overrides": {"Producto": pl.Utf8}}),
        ("importados", "importados.csv", {"schema_overrides": S}),
    ):
        df = _snapshot(archivo, **kw)
        if df is not None:
            fuentes[nombre] = df

    # --- Seguimiento de compras ---
    hay_seguimiento_excel = any(
        x is not None
        for x in (seguimiento_nacional_xlsx, seguimiento_importado_xlsx, seguimiento_frontera_xlsx)
    )
    if hay_seguimiento_excel:
        fuentes.update(
            _seguimiento_desde_excel(
                seguimiento_nacional_xlsx, seguimiento_importado_xlsx, seguimiento_frontera_xlsx
            )
        )
    elif sql_conn is not None:
        from .conectores import sql_flexline

        fuentes["seguimiento"] = sql_flexline.leer_seguimiento(sql_conn, para_transito=False)
        fuentes["seguimiento_transito"] = sql_flexline.leer_seguimiento(sql_conn, para_transito=True)
        fuentes["seguimiento_lt"] = sql_flexline.leer_seguimiento_lt(sql_conn)
    else:
        for nombre in ("seguimiento", "seguimiento_transito", "seguimiento_lt"):
            fuentes[nombre] = pl.read_csv(
                snap / f"{nombre}.csv", schema_overrides=S, try_parse_dates=True
            )

    # --- Ventas ---
    ventas_crudo: pl.DataFrame | None = None
    if ventas_xlsx is not None:
        from . import lectores_excel as lx
        from .conectores import sql_flexline as sf

        ventas_crudo = lx.leer_ventas_excel(ventas_xlsx)
        # 'Ventas Unificadas' del modelo = Curifor (E01) UNION Frontera (E07). Sin la
        # parte de Frontera el motor pierde los combos que solo venden ahi y subestima
        # la demanda de los que venden en las dos.
        fuentes["ventas"] = sf.unir_ventas(
            sf.normalizar_ventas_curifor(ventas_crudo),
            sf.normalizar_ventas_frontera(ventas_frontera_crudo)
            if ventas_frontera_crudo is not None
            else None,
        )
    elif sql_conn is not None:
        from .conectores import sql_flexline

        fuentes["ventas"] = sql_flexline.leer_ventas_curifor(sql_conn)
    else:
        fuentes["ventas"] = pl.read_csv(
            snap / "ventas_12m.csv", schema_overrides=S, try_parse_dates=True
        )

    # --- Tablas chicas calculadas (reemplazan el snapshot congelado del BI) ---
    # Cada una se calcula solo si estan sus insumos; si falta alguno, se queda el
    # snapshot. Asi se puede cortar el cordon de a una tabla, sin quedar a medias.
    if seguimiento_importado_xlsx is not None:
        from . import lectores_excel as lx

        fuentes["importados"] = dim.calcular_importados(
            lx.leer_seguimiento_importado_excel(seguimiento_importado_xlsx)
        )

    if listado_maestro is not None and ventas_crudo is not None:
        from . import lectores_excel as lx

        # Productos "de Frontera" para la regla CHEVROLET de FILTRO1_Final.
        en_frontera = pl.concat([
            stock_frontera.select("Producto"),
            (
                ventas_frontera_crudo.select(pl.col("producto").alias("Producto"))
                if ventas_frontera_crudo is not None
                else pl.DataFrame(schema={"Producto": pl.Utf8})
            ),
        ]).get_column("Producto").drop_nulls().unique().to_list()

        dp = dim.calcular_dim_producto(
            ventas_crudo, lx.leer_listado_maestro(listado_maestro), stock, en_frontera
        )
        fuentes["dim_producto"] = dp.select(["Producto", "Categoria"])
        fuentes["catalogo"] = dp.select(
            ["Producto", "Descripcion", "FILTRO1_Final", "UnidadMedida"]
        )

    if mix_reemplazos_xlsx is not None and ventas_crudo is not None and fin_mes_cerrado:
        from . import lectores_excel as lx

        fuentes["mapeo"] = dim.calcular_mapeo_master(
            lx.leer_mix_reemplazos(mix_reemplazos_xlsx), ventas_crudo, fin_mes_cerrado
        )

    return fuentes
