"""Vigentes de FORD consultados en vivo (proyecto WINGS, `lista_new*.xlsx`).

La lista de precios es una foto y no trae los codigos que FORD dio de baja: medido
sobre los 33 repuestos FORD de la pauta InStock (18-08-2026), 9 tienen reemplazo y
8 de esos 9 NO aparecen en la lista. WINGS consulta el portal y si los resuelve.

Las dos fuentes se COMBINAN, no se pisan: WINGS manda hacia adelante (quien
reemplaza a este codigo) y la lista de precios manda en la inversa (`reemplaza_a`),
que WINGS no publica. De los 33 codigos de la pauta, 11 traen `reemplaza_a` en la
lista; pisarlos enteros perderia esos 11 grupos sin que nada avisara.
"""
from datetime import date

import polars as pl
from openpyxl import Workbook

from src.motor.dimensiones import ampliar_mapeo_con_ford
from src.motor.lectores_excel import (
    ESQUEMA_REEMPLAZOS_FORD,
    combinar_reemplazos_ford,
    leer_reemplazos_ford,
    leer_vigentes_ford,
)

CABECERAS = [
    "Codigo_Original", "Origen", "Codigo_POPIMS", "Encontrado", "Vigente",
    "Tiene_Reemplazo", "Stock_Ford", "Precio_Promedio_Vigente", "Aviso",
]


def _xlsx(ruta, filas, hoja="lista_new"):
    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    ws.append(CABECERAS)
    for f in filas:
        ws.append(f)
    wb.save(ruta)


def _fila(codigo, popims, *, encontrado=True, vigente=None, aviso=None):
    """Una fila de la salida de WINGS, en el orden de `CABECERAS`."""
    return [codigo, "Personalizada", popims, encontrado, vigente or "",
            bool(vigente), True, None, aviso]


def _reem(**kw):
    base = {
        "clave_precio": None, "sku_ford": None, "clave_vigente": None,
        "sku_vigente": None, "cadena": None, "reemplaza_a": [],
        "estado_reemplazo": None, "sucesor_confirmado": False, "aviso": None,
    }
    base.update(kw)
    return base


def _frame(filas):
    return pl.DataFrame(filas, schema=dict(ESQUEMA_REEMPLAZOS_FORD))


VENTAS_VACIAS = pl.DataFrame(
    {"Producto": [], "Fecha": [], "Cantidad": []},
    schema={"Producto": pl.Utf8, "Fecha": pl.Date, "Cantidad": pl.Float64},
)
FIN = date(2026, 8, 1)


# --- El lector -------------------------------------------------------------------

def test_lee_el_vigente_y_arma_la_cadena(tmp_path):
    ruta = tmp_path / "lista_new.xlsx"
    _xlsx(ruta, [
        _fila("MB3Z19N619C", "MB3Z/19N619/C/", vigente="MB3Z/19N619/A/"),
    ])
    df = leer_vigentes_ford(ruta)

    assert df.height == 1
    f = df.to_dicts()[0]
    assert f["clave_precio"] == "MB3Z19N619C"
    assert f["clave_vigente"] == "MB3Z19N619A"
    assert f["sku_vigente"] == "MB3Z/19N619/A/"
    # Mismo separador que la lista de precios: la plataforma pinta la cadena tal
    # cual, y con otro formato se veria distinta segun de que fuente vino.
    assert f["cadena"] == "MB3Z/19N619/C/ > MB3Z/19N619/A/"
    assert f["sucesor_confirmado"] is True


def test_las_dos_fuentes_devuelven_el_mismo_esquema(tmp_path):
    """Es lo que hace posible combinarlas. Si una se desalinea, la mezcla falla en
    silencio: el motor pierde reemplazos y ningun test revienta."""
    ruta_w = tmp_path / "lista_new.xlsx"
    _xlsx(ruta_w, [_fila("MB3Z19N619C", "MB3Z/19N619/C/", vigente="MB3Z/19N619/A/")])

    ruta_p = tmp_path / "precios.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Precios"
    ws.append(["PartNumber", "Reemplazado_Por", "Cadena_Reemplazo", "Reemplaza_A",
               "Estado_Reemplazo", "Reemplazo_Aviso"])
    ws.append(["KK3Z/3504/BR/", "KK3Z/3504/U/", None, None, "Encontrado", None])
    wb.save(ruta_p)

    assert leer_vigentes_ford(ruta_w).schema == leer_reemplazos_ford(ruta_p).schema
    assert list(leer_vigentes_ford(ruta_w).schema) == list(ESQUEMA_REEMPLAZOS_FORD)


def test_sin_reemplazo_no_deja_vigente(tmp_path):
    ruta = tmp_path / "lista_new.xlsx"
    _xlsx(ruta, [_fila("GK2Z9601B", "GK2Z/9601/B/")])
    f = leer_vigentes_ford(ruta).to_dicts()[0]

    assert f["clave_vigente"] is None
    assert f["sku_vigente"] is None
    assert f["cadena"] is None
    assert f["sucesor_confirmado"] is False


def test_la_fila_sin_respuesta_de_ford_queda_fuera(tmp_path):
    """"No se sabe" no es lo mismo que "no tiene reemplazo".

    Cuando WINGS no pudo traducir el codigo al formato POPIMS, FORD nunca lo
    respondio. Si esa fila entrara como "sin reemplazo", al combinar borraria el
    sucesor que la lista de precios si trae para ese codigo. En la corrida del
    18-08-2026 es 1 de 33 (`CYFS12YRT3`), y en la lista tiene dos reemplazos.
    """
    ruta = tmp_path / "lista_new.xlsx"
    _xlsx(ruta, [
        _fila("CYFS12YRT3", "", encontrado=None,
              aviso="no se pudo determinar el formato POPIMS: revisar a mano"),
        _fila("GK2Z9601B", "GK2Z/9601/B/"),
    ])
    df = leer_vigentes_ford(ruta)

    assert df["clave_precio"].to_list() == ["GK2Z9601B"]


def test_acepta_los_booleanos_como_texto(tmp_path):
    """Un Excel reabierto y guardado a mano deja TRUE/VERDADERO en vez de bool."""
    ruta = tmp_path / "lista_new.xlsx"
    _xlsx(ruta, [
        ["MB3Z19N619C", "Personalizada", "MB3Z/19N619/C/", "VERDADERO",
         "MB3Z/19N619/A/", "TRUE", "TRUE", None, None],
    ])
    f = leer_vigentes_ford(ruta).to_dicts()[0]

    assert f["clave_vigente"] == "MB3Z19N619A"
    assert f["sucesor_confirmado"] is True


def test_hoja_con_otras_columnas_no_revienta(tmp_path):
    ruta = tmp_path / "cualquiera.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "lista_new"
    ws.append(["Codigo", "Otra"])
    ws.append(["X", 1])
    wb.save(ruta)

    df = leer_vigentes_ford(ruta)
    assert df.is_empty()
    assert list(df.schema) == list(ESQUEMA_REEMPLAZOS_FORD)


# --- La combinacion --------------------------------------------------------------

def test_conserva_el_reemplaza_a_de_la_lista_de_precios():
    """La regresion que este test existe para atajar.

    WINGS resuelve hacia adelante y devuelve `reemplaza_a` vacio. Pisar la fila
    entera se veria mas simple y perderia la direccion inversa, que es de donde
    sale casi todo lo que Curifor efectivamente agrupa.
    """
    lista = _frame([_reem(clave_precio="BK3Z9601B",
                          reemplaza_a=["1945831", "1881895"])])
    wings = _frame([_reem(clave_precio="BK3Z9601B", sku_ford="BK3Z/9601/B/")])

    out = combinar_reemplazos_ford(lista, wings)

    assert out.height == 1
    assert out.to_dicts()[0]["reemplaza_a"] == ["1945831", "1881895"]


def test_wings_manda_hacia_adelante():
    lista = _frame([_reem(clave_precio="MB3Z19N619C", clave_vigente="OTRO",
                          sucesor_confirmado=False)])
    wings = _frame([_reem(clave_precio="MB3Z19N619C", clave_vigente="MB3Z19N619A",
                          sku_vigente="MB3Z/19N619/A/", sucesor_confirmado=True)])

    f = combinar_reemplazos_ford(lista, wings).to_dicts()[0]

    assert f["clave_vigente"] == "MB3Z19N619A"
    assert f["sucesor_confirmado"] is True


def test_wings_borra_un_sucesor_que_la_lista_declaraba_de_mas():
    """Caso real: la lista de precios daba `7C3Z9601A -> 7C3Z9601C` con estado
    "Sin candidato vigente"; WINGS consulto y el codigo esta vigente. Gana WINGS,
    que miro el portal hoy, pero la direccion inversa se conserva igual."""
    lista = _frame([_reem(clave_precio="7C3Z9601A", clave_vigente="7C3Z9601C",
                          estado_reemplazo="Sin candidato vigente",
                          reemplaza_a=["AL3Z9601A"])])
    wings = _frame([_reem(clave_precio="7C3Z9601A", sku_ford="7C3Z/9601/A/")])

    f = combinar_reemplazos_ford(lista, wings).to_dicts()[0]

    assert f["clave_vigente"] is None
    assert f["reemplaza_a"] == ["AL3Z9601A"]


def test_las_filas_que_wings_no_toca_quedan_intactas():
    lista = _frame([
        _reem(clave_precio="AAA", clave_vigente="BBB", sucesor_confirmado=True),
        _reem(clave_precio="CCC", reemplaza_a=["DDD"]),
    ])
    wings = _frame([_reem(clave_precio="CCC", sku_ford="CCC")])

    out = combinar_reemplazos_ford(lista, wings).sort("clave_precio")

    assert out.height == 2
    aaa = out.filter(pl.col("clave_precio") == "AAA").to_dicts()[0]
    assert aaa["clave_vigente"] == "BBB" and aaa["sucesor_confirmado"] is True


def test_un_codigo_que_solo_conoce_wings_se_agrega():
    lista = _frame([_reem(clave_precio="AAA")])
    wings = _frame([_reem(clave_precio="BR3Z8620S", clave_vigente="RB5Z8620D",
                          sucesor_confirmado=True)])

    out = combinar_reemplazos_ford(lista, wings)

    assert sorted(out["clave_precio"].to_list()) == ["AAA", "BR3Z8620S"]
    nuevo = out.filter(pl.col("clave_precio") == "BR3Z8620S").to_dicts()[0]
    assert nuevo["reemplaza_a"] == []


def test_sin_wings_la_lista_queda_igual():
    """Es el camino de todos los dias hasta que el archivo de WINGS exista."""
    lista = _frame([_reem(clave_precio="AAA", clave_vigente="BBB")])

    assert combinar_reemplazos_ford(lista, _frame([])).equals(lista)


def test_sin_lista_de_precios_sirve_wings_solo():
    wings = _frame([_reem(clave_precio="AAA", clave_vigente="BBB")])

    out = combinar_reemplazos_ford(_frame([]), wings)

    assert out.height == 1
    assert list(out.schema) == list(ESQUEMA_REEMPLAZOS_FORD)


# --- De punta a punta: que el par termine agrupado --------------------------------

def _mapeo(pares=()):
    return pl.DataFrame(
        [{"Producto": p, "Producto_Master": m} for p, m in pares],
        schema={"Producto": pl.Utf8, "Producto_Master": pl.Utf8},
    )


def test_el_par_de_wings_queda_bajo_el_codigo_vigente():
    """El caso que reporto Abastecimiento: la plataforma pedia `25 MB3Z19N619C`,
    dado de baja, teniendo `19 MB3Z19N619A` en el catalogo."""
    wings = _frame([_reem(clave_precio="MB3Z19N619C", clave_vigente="MB3Z19N619A",
                          sku_vigente="MB3Z/19N619/A/", sucesor_confirmado=True)])
    conocidos = ["25 MB3Z19N619C", "19 MB3Z19N619A"]

    out = ampliar_mapeo_con_ford(_mapeo(), wings, conocidos, VENTAS_VACIAS, FIN)

    masters = dict(out.select(["Producto", "Producto_Master"]).iter_rows())
    assert masters["25 MB3Z19N619C"] == "19 MB3Z19N619A"
    assert masters["19 MB3Z19N619A"] == "19 MB3Z19N619A"


def test_si_el_vigente_no_esta_en_curifor_no_se_agrupa_nada():
    """`25 MB3Z8620E` -> `MB3Z/8620/F/`, que no esta en el maestro (medido
    22-08-2026, 1 de los 9 pares de la pauta). El ERP no puede comprar un codigo
    que no conoce, asi que el viejo se queda como esta y el aviso lo da la
    plataforma con `reemplazado_por_ford`. No hay que programar el fallback: sale
    solo de que `por_clave` no encuentra el vigente."""
    wings = _frame([_reem(clave_precio="MB3Z8620E", clave_vigente="MB3Z8620F",
                          sku_vigente="MB3Z/8620/F/", sucesor_confirmado=True)])

    out = ampliar_mapeo_con_ford(_mapeo(), wings, ["25 MB3Z8620E"], VENTAS_VACIAS, FIN)

    assert out.is_empty()


def test_el_resultado_no_depende_del_orden_de_los_productos():
    """El motor tiene que dar lo mismo dos veces sobre la misma base.

    2.399 claves del maestro tienen mas de un codigo de Curifor: "25 CN1Z8620E" y
    "19 CN1Z8620E" son el mismo repuesto con distinto rubro y normalizan igual.
    `por_clave` se queda con uno, y hasta el 22-08-2026 ese uno dependia del orden
    en que llegaran los productos —que sale de un `.unique()` de polars y no es
    estable—. Dos corridas identicas daban 15 y 19 filas distintas, y un producto
    pedia $61.286 en una y $137.152 en la otra.
    """
    wings = _frame([_reem(clave_precio="CN1Z8620D", clave_vigente="CN1Z8620E",
                          sucesor_confirmado=True)])
    duplicados = ["25 CN1Z8620E", "19 CN1Z8620E", "15 CN1Z8620E", "25 CN1Z8620D"]

    resultados = [
        dict(
            ampliar_mapeo_con_ford(_mapeo(), wings, orden, VENTAS_VACIAS, FIN)
            .select(["Producto", "Producto_Master"]).iter_rows()
        )
        for orden in (duplicados, list(reversed(duplicados)), sorted(duplicados))
    ]

    assert resultados[0] == resultados[1] == resultados[2]
    assert resultados[0]["25 CN1Z8620D"] == "15 CN1Z8620E"


def test_el_mix_sigue_mandando_sobre_wings():
    """Misma regla que con la lista de precios: un producto ya agrupado por el mix
    no lo toca FORD. Se midio el 07-08-2026 que invertirlo rompia 41 grupos."""
    wings = _frame([_reem(clave_precio="MB3Z19N619C", clave_vigente="MB3Z19N619A",
                          sucesor_confirmado=True)])
    mix = _mapeo([("25 MB3Z19N619C", "OTRO MASTER"), ("OTRO MASTER", "OTRO MASTER")])

    out = ampliar_mapeo_con_ford(
        mix, wings, ["25 MB3Z19N619C", "19 MB3Z19N619A", "OTRO MASTER"],
        VENTAS_VACIAS, FIN,
    )

    masters = dict(out.select(["Producto", "Producto_Master"]).iter_rows())
    assert masters["25 MB3Z19N619C"] == "OTRO MASTER"
